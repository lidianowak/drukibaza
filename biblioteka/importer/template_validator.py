from pathlib import Path

from openpyxl import load_workbook


HEADER_ROW = 14


def validate_template_structure(
    workbook,
    result,
):

    """
    Sprawdza zgodność struktury formularza
    z aktualnym formularzem wzorcowym.
    """

    template_path = (
        Path(__file__).parent
        / "resources"
        / "formularz importu_BiDO_pusty.xlsx"
    )

    template = load_workbook(template_path)

    for sheet_name in template.sheetnames:

        if sheet_name not in workbook.sheetnames:
            result.add_error(
                message=f"Brakuje arkusza '{sheet_name}'.",
                sheet="Plik",
            )
            continue

        expected = [
            cell.value
            for cell in template[sheet_name][HEADER_ROW]
        ]

        actual = [
            cell.value
            for cell in workbook[sheet_name][HEADER_ROW]
        ]

        # usuń puste komórki z końca wiersza
        while expected and expected[-1] is None:
            expected.pop()

        while actual and actual[-1] is None:
            actual.pop()

        max_columns = max(len(expected), len(actual))

        for i in range(max_columns):

            expected_name = (
                expected[i]
                if i < len(expected)
                else None
            )

            actual_name = (
                actual[i]
                if i < len(actual)
                else None
            )

            if expected_name == actual_name:
                continue

            if expected_name is None:
                result.add_error(
                    message=(
                        f"Nieoczekiwana kolumna "
                        f"'{actual_name}' "
                        f"(kolumna {i+1})."
                    ),
                    sheet=sheet_name,
                )

            elif actual_name is None:
                result.add_error(
                    message=(
                        f"Brakuje kolumny "
                        f"'{expected_name}' "
                        f"(kolumna {i+1})."
                    ),
                    sheet=sheet_name,
                )

            else:
                result.add_error(
                    message=(
                        f"Kolumna {i+1}: "
                        f"oczekiwano '{expected_name}', "
                        f"otrzymano '{actual_name}'."
                    ),
                    sheet=sheet_name,
                )