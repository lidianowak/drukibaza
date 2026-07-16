"""
template_generator.py

Generowanie formularza importu BiDO.
"""

from pathlib import Path

from openpyxl import load_workbook

from biblioteka.models import (
    Jezyk,
    Format,
    Czcionka,
)


def load_template():
    """
    Wczytuje wzorcowy formularz importu.
    """

    path = (
        Path(__file__)
        .parent
        / "resources"
        / "formularz importu_BiDO_pusty.xlsx"
    )

    workbook = load_workbook(path)

    return workbook

def update_dictionary(
    workbook,
    model,
    column,
    title,
):
    """
    Aktualizuje kolumnę słownika w arkuszu _Słowniki.
    """

    sheet = workbook["_Słowniki"]

    sheet[f"{column}1"] = title

    row = 2

    for obj in model.objects.order_by("nazwa"):

        sheet.cell(
            row=row,
            column=ord(column) - 64,
            value=obj.nazwa,
        )

        row += 1


def update_dictionaries(workbook):
    """
    Aktualizuje wszystkie słowniki formularza.
    """

    update_dictionary(
        workbook,
        Jezyk,
        "A",
        "Języki",
    )

    update_dictionary(
        workbook,
        Format,
        "B",
        "Formaty",
    )

    update_dictionary(
        workbook,
        Czcionka,
        "C",
        "Czcionki",
    )

def generate_template():
    """
    Generuje aktualny formularz importu.
    """

    workbook = load_template()

    update_dictionaries(workbook)

    return workbook