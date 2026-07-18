from pathlib import Path

from openpyxl import load_workbook as openpyxl_load_workbook


def load_workbook(path: Path):
    """
    Otwiera formularz importu BiDO.
    """

    return openpyxl_load_workbook(path)